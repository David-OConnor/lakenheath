from collections import namedtuple

import arrow
import openpyxl as px

duty = namedtuple('Duty', ['description', 'notes', 'names', 'time'])

meeting = namedtuple('Meeting', ['time', 'description', 'location'])

flight = namedtuple('Flight', ['show_date', 'line_num', 'push_time', 'callsign',
                               'pilot1', 'wso1', 'pilot2', 'wso2', 'remarks'])

# Draws a rectanble from the top left cell to bottom right cell.
section_ranges = {
    'duties': (('C', 6), ('E', 17)),
    'meetings': (('M', 7), ('O', 14)),
    'flying': (('K', 17), ('O', 42)),
    'dnif': (('C', 21), ('E', 21)),
    'unavailable': (('C', 22), ('E', 22)),
    'not scheduled': (('B', 25), ('E', 45))
}


def letter_range(start, end):
    """Like range, but for letters instead of numbers."""
    return (chr(item) for item in range(ord(start), ord(end)+1))


def load_workbook(filename):
    W = px.load_workbook(filename, use_iterators=True)
    p = W.get_sheet_by_name(name='Schedule')
    return p


def find_section(col, row):
    for sec, sec_range in section_ranges.items():
        col_range = letter_range(sec_range[0][0], sec_range[1][0])
        row_range = range(sec_range[0][1], sec_range[1][1] + 1)

        if col in col_range and row in row_range:
            return sec


def cell_from_name(name, p):
    # todo be flexible with name format
    for row in p.iter_rows():
        for cell in row:
            if cell.value is None or type(cell.value) == float:
                continue
            if name.upper() in cell.value.upper():
                yield cell.column, cell.row, cell.value


def find_shift(p, row, test_col):
    """Find how many cells to shift up when encountering possible merged cells."""
    max_shift = 4
    for shift in range(max_shift):
        test_val = p[test_col + str(row - shift)].value

        # If None,the cells were merged vertically; look up.
        if test_val is not None:
            return shift
        shift += 1
    raise AttributeError("Problem parsing merged cell.")


def helper(start_col, end_col, row, p, shift_col=0):
    """Reduces repetition in personal_schedule(), at the cost of flexibility."""
    cols = list(letter_range(start_col, end_col))
    shift = find_shift(p, row, cols[shift_col])
    for col in cols:
        yield p[col + str(row - shift)].value


def personal_schedule(name, p):
    result = []

    cells = cell_from_name(name, p)
    if not cells:
        return

    for col, row, value in cells:
        section = find_section(col, row)

        if section == 'duties':
            row -= find_shift(p, row, 'B')

            description = p['B' + str(row)].value
            notes = p['C' + str(row)].value
            names = p['D' + str(row)].value
            time = p['E' + str(row)].value
            result.append(duty(description, notes, names, time))

            # fields = helper('B', 'E', row, p)
            # result.append(duty(*fields))

        elif section == 'meetings':
            row -= find_shift(p, row, 'G')
            
            time = p['G' + str(row)].value
            description = p['H' + str(row)].value
            location = p['L' + str(row)].value
            result.append(meeting(time, description, location))
            #
            # fields = helper('G', 'L', row, p)
            # result.append(meeting(*fields))

        elif section == 'flying':
            fl_row = row - find_shift(p, row, 'H')

            show_date = p['G' + str(fl_row)].value
            line_num = p['H' + str(fl_row)].value
            push_time = p['I' + str(fl_row)].value
            callsign = p['J' + str(fl_row)].value

            pilot_1 = p['K' + str(fl_row)].value
            wso_1 = p['L' + str(fl_row)].value
            pilot_2 = p['K' + str(fl_row + 1)].value
            wso_2 = p['L' + str(fl_row + 1)].value

            remarks = p['M' + str(fl_row)].value
            result.append(flight(show_date, line_num, push_time, callsign,
                                 pilot_1, wso_1, pilot_2, wso_2, remarks))

        elif section == 'dnif':
            result.append("DNIF; sorry bro.")

        elif section == 'unavailable':
            result.append("Unavailable.")

        elif section == 'not scheduled':
            result.append("Not scheduled.")

        else:
            raise AttributeError("Not sure what you're scheduled for.")

    for item in result:
        print(item)

    # return result
pass